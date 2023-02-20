# -*- coding: utf-8 -*-
"""
Created on Thu Nov 24 13:53:02 2022

@author: NM DUC
"""
import numpy as np
from selenium import webdriver
from time import sleep
import random
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
from selenium.webdriver.common.by import By
import pandas as pd
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
import json
import urllib3
import json
import urllib.request



excel = pd.read_excel('./auto_find.xlsx')

options = webdriver.ChromeOptions()
options.headless = True
# OR options.add_argument("--disable-gpu")


# Declare browser
driver = webdriver.Chrome(executable_path='./chromedriver.exe',chrome_options=options)

# lấy dữ liệu trong file excel đầu vào 
links = excel['Nhập url'].values.tolist()
#text_finds = excel['Từ khóa cần tìm'].values.tolist()

count_text = []
urls = []
texts = []


# Open URL
for i in range(0,len(links)):
    
    #lấy url ra để xét    
    url = str(links[i]) 
    
    #lấy text cần tìm trong url ở trên
    #text = str(text_finds[i])
    text = 'AMIS CRM'
    try:
        print(url)
        req = requests.get(url)
        #req = requests.get('https://asp.misa.vn/')
        print(req.status_code)
    except requests.exceptions.RequestException as e:  # This is the correct syntax
        raise SystemExit(e)

    #lấy toàn bộ nội dung website
    soup = BeautifulSoup(req.text, "lxml")
    
    # đưa url đang duyệt vào danh sách url đã duyệt 
    urls.append(url)
    #sleep(1)
    # crawler = driver.find_elements(By.CSS_SELECTOR, ".td-post-content.tagdiv-type")
    #crawler = driver.find_elements(By.CSS_SELECTOR, "head")
    # bắn js tìm
    
    
    # Check từ đấy có trong bài không
    
    #tìm tiêu đề và nội dung
    #titleFind = soup.find_all('div', class_='current-title')[0].text
    #contentFind = soup.find_all('div', class_='td-post-content tagdiv-type')[0].text
    #allTextFind = titleFind + contentFind
    
    #lấy chuẩn hơn
    #lọc lấy hết thẻ p trong td-post-content tagdiv-type (có vẻ có cả heading)
    article_text = ''
    article = soup.find("div", {"class":"td-post-content tagdiv-type"}).find_all('p')
    for element in article:
        article_text += '\n' + ''.join(element.findAll(text = True))
        
        
    #lấy tiêu đề
    titleFind = soup.find_all('div', class_='current-title')[0].text
    
    #tổng text
    allTextFind = titleFind + article_text
    
    #check xem số lượng bao nhiêu
    findCountText = allTextFind.lower().count(text.lower())
    
    #tìm xong thì đưa vào text
    texts.append(findCountText)

print(urls)
df1 = pd.DataFrame(list(zip(urls, count_text)), columns = ['title', 'số lần hiện'])
print(df1)


output_excel_path= 'ket_qua.xlsx'



#Xác định số hàng và cột lớn nhất trong file excel cần tạo
row = len(urls)
column = len(str(texts))

#Tạo một workbook mới và active nó
wb = openpyxl.Workbook()
sheet = wb.active


# (hàng,cột)
sheet.cell(1,1, value= 'Nhập url')
for a in range(0,row):
    sheet.cell(a + 2,1, value= links[a])

    
sheet.cell(1,2, value= 'Số từ xuất hiện')
for i in range(0,row):
    sheet.cell(i + 2 , 2, value= texts[i])
    

    
   
#Lưu lại file Excel
wb.save(output_excel_path)

# Close browser
driver.close()   
