# -*- coding: utf-8 -*-
"""
Created on Thu Nov 24 13:53:02 2022

@author: NM DUC
"""
import numpy as np
from selenium import webdriver
from time import sleep
import random
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
from selenium.webdriver.common.by import By
import pandas as pd
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
import json
import urllib3
import json
import urllib.request



excel = pd.read_excel('./auto_find.xlsx')

options = webdriver.ChromeOptions()
options.headless = True
# OR options.add_argument("--disable-gpu")


# Declare browser
driver = webdriver.Chrome(executable_path='./chromedriver.exe',chrome_options=options)

# lấy dữ liệu trong file excel đầu vào 
links = excel['Nhập url'].values.tolist()
text_finds = excel['Từ khóa cần tìm'].values.tolist()

count_text = []
urls = []
texts = []


# Open URL
for i in range(0,len(links)):
    # url = "https://amis.misa.vn" +'/'+ str(link) + '/'
    url = str(links[i]) 
    try:
        print(url)
        req = requests.get(url)
        #req = requests.get('https://asp.misa.vn/')
        print(req.status_code)
    except requests.exceptions.RequestException as e:  # This is the correct syntax
        raise SystemExit(e)

    #lấy toàn bộ nội dung website
    soup = BeautifulSoup(req.text, "lxml")
    
    # đưa url đang duyệt vào danh sách url đã duyệt 
    urls.append(url)
    #sleep(2)
    # crawler = driver.find_elements(By.CSS_SELECTOR, ".td-post-content.tagdiv-type")
    #crawler = driver.find_elements(By.CSS_SELECTOR, "head")
    # bắn js tìm
    
    
    crawler = soup.find_all("a", {"class": "entry-crumb"})[-1].text
    
    text = str(crawler)
    texts.append(text)

print(urls)
df1 = pd.DataFrame(list(zip(urls, count_text)), columns = ['title', 'số lần hiện'])
print(df1)


output_excel_path= 'ket_qua.xlsx'



#Xác định số hàng và cột lớn nhất trong file excel cần tạo
row = len(urls)
column = len(texts[0])

#Tạo một workbook mới và active nó
wb = openpyxl.Workbook()
sheet = wb.active


# (hàng,cột)
sheet.cell(1,1, value= 'Nhập url')
for a in range(0,row):
    sheet.cell(a + 2,1, value= links[a])

sheet.cell(1,2, value= 'Từ khóa cần tìm')
for i in range(0,row):
    sheet.cell(i + 2 , 2, value= texts[i])
    

    
   
#Lưu lại file Excel
wb.save(output_excel_path)

# Close browser
driver.close()   
